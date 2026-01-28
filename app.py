#!/usr/bin/env python3
"""Streamlit app to extract 'ПРЕДАЛ' and build a filtered table from an Excel file."""

import io
import time
from datetime import date, datetime
from html.parser import HTMLParser
from typing import Dict, List, Optional, Tuple

import requests
import streamlit as st
from openpyxl import Workbook, load_workbook


UA = "Mozilla/5.0 (compatible; lagardere_app/1.0)"


class TextExtractor(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.chunks: List[str] = []

    def handle_data(self, data: str) -> None:
        data = data.strip()
        if data:
            self.chunks.append(data)


def html_to_chunks(html: str) -> List[str]:
    parser = TextExtractor()
    try:
        parser.feed(html)
        parser.close()
    except Exception:
        pass
    return parser.chunks


def extract_predal_from_chunks(chunks: List[str]) -> Optional[str]:
    upper_chunks = [c.upper() for c in chunks]
    for i, chunk_upper in enumerate(upper_chunks):
        if "ПРЕДАЛ" in chunk_upper:
            original = chunks[i]
            idx = original.upper().find("ПРЕДАЛ")
            after = original[idx + len("ПРЕДАЛ"):]
            after = after.lstrip(" :\u00a0")
            if after:
                return after
            if i + 1 < len(chunks):
                next_chunk = chunks[i + 1].strip()
                if next_chunk and not next_chunk.endswith(":"):
                    return next_chunk
    return None


def fetch_html(url: str, timeout: float) -> Tuple[str, str]:
    headers = {"User-Agent": UA}
    resp = requests.get(url, headers=headers, timeout=timeout)
    resp.raise_for_status()
    content_type = resp.headers.get("Content-Type", "")
    resp.encoding = resp.apparent_encoding or resp.encoding
    return resp.text, content_type


def build_header_map(ws) -> Dict[str, int]:
    header_map: Dict[str, int] = {}
    for cell in ws[1]:
        if cell.value is None:
            continue
        header_map[str(cell.value).strip()] = cell.column
    return header_map


def format_cell(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    return str(value).strip()


def load_rows(
    file_bytes: bytes,
    sheet: Optional[str],
    number_header: str,
    client_header: str,
    date_header: str,
    product_header: str,
    quantity_header: str,
    client_prefix: str,
) -> List[Tuple[str, str, str, str, str, Optional[str]]]:
    wb = load_workbook(io.BytesIO(file_bytes))
    ws = wb[sheet] if sheet else wb.active

    header_map = build_header_map(ws)

    def col(name: str, default: int) -> int:
        return header_map.get(name, default)

    number_col = col(number_header, 2)
    client_col = col(client_header, 4)
    date_col = col(date_header, 6)
    product_col = col(product_header, 8)
    quantity_col = col(quantity_header, 10)

    rows: List[Tuple[str, str, str, str, str, Optional[str]]] = []
    prefix = client_prefix.strip().lower()

    for r in range(2, ws.max_row + 1):
        client = ws.cell(row=r, column=client_col).value
        if client is None:
            continue
        client_str = str(client).strip()
        if not client_str.lower().startswith(prefix):
            continue

        number_cell = ws.cell(row=r, column=number_col)
        number = number_cell.value
        if number is None:
            continue
        number_str = format_cell(number)

        date_str = format_cell(ws.cell(row=r, column=date_col).value)
        product_str = format_cell(ws.cell(row=r, column=product_col).value)
        qty_str = format_cell(ws.cell(row=r, column=quantity_col).value)

        link = None
        if number_cell.hyperlink and number_cell.hyperlink.target:
            link = number_cell.hyperlink.target
        elif isinstance(number_cell.value, str) and number_cell.value.startswith("http"):
            link = number_cell.value

        rows.append((number_str, client_str, date_str, product_str, qty_str, link))

    return rows


def build_output(rows: List[Tuple[str, str, str, str, str]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Lagardere"
    ws.append(["Номер", "Предал", "Дата", "Продукт", "Количество"])
    for row in rows:
        ws.append(list(row))
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def main() -> None:
    st.set_page_config(page_title="Лагардер – Стокови разписки", layout="wide")
    st.title("Извличане на стокови разписки (Лагардер)")

    uploaded = st.file_uploader("Качи Excel (.xlsx)", type=["xlsx"])
    client_prefix = st.text_input("Фирма започва с", "Лагардер")
    fetch_predal = st.checkbox("Извличай 'ПРЕДАЛ' от хиперлинковете", value=True)
    delay = st.number_input("Пауза между заявки (сек)", min_value=0.0, max_value=5.0, value=0.0, step=0.1)
    timeout = st.number_input("Timeout (сек)", min_value=5.0, max_value=60.0, value=20.0, step=1.0)

    if not uploaded:
        st.info("Качи файл, за да започнем.")
        return

    file_bytes = uploaded.read()
    try:
        rows = load_rows(
            file_bytes,
            sheet=None,
            number_header="Номер",
            client_header="Клиент",
            date_header="Дата на документа",
            product_header="Наименование на продукта",
            quantity_header="Количество",
            client_prefix=client_prefix,
        )
    except Exception as exc:
        st.error(f"Грешка при четене на Excel: {exc}")
        return

    if not rows:
        st.warning("Няма редове за този клиент.")
        return

    url_cache: Dict[str, Optional[str]] = {}
    output_rows: List[Tuple[str, str, str, str, str]] = []

    progress = st.progress(0.0)
    status = st.empty()

    total = len(rows)
    for i, (number, _client, date_str, product_str, qty_str, link) in enumerate(rows, start=1):
        predal = ""
        if fetch_predal and link:
            if link in url_cache:
                predal = url_cache[link] or ""
            else:
                try:
                    html, content_type = fetch_html(link, float(timeout))
                    if "text/html" in content_type:
                        chunks = html_to_chunks(html)
                        predal = extract_predal_from_chunks(chunks) or ""
                    url_cache[link] = predal
                except Exception:
                    url_cache[link] = ""
            if delay:
                time.sleep(float(delay))

        output_rows.append((number, predal, date_str, product_str, qty_str))
        progress.progress(i / total)
        status.text(f"Обработени: {i}/{total}")

    st.subheader("Преглед")
    st.dataframe(output_rows, use_container_width=True)

    out_bytes = build_output(output_rows)
    st.download_button(
        "Свали резултата (.xlsx)",
        data=out_bytes,
        file_name="lagardere_output.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


if __name__ == "__main__":
    main()
