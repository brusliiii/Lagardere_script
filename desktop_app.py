#!/usr/bin/env python3
"""Desktop app (PyQt6) for extracting 'ПРЕДАЛ' and building a filtered table."""

import os
import time
from datetime import date, datetime
from html.parser import HTMLParser
from typing import Dict, List, Optional, Tuple

import requests
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from PyQt6 import QtCore, QtGui, QtWidgets

UA = "Mozilla/5.0 (compatible; lagardere_desktop/1.0)"

PRODUCT_ORDER = [
    "PAZ X-Freeze",
    "PAZ X-Freeze +",
    "PAZ Cool Mint",
    "PAZ Cool mint +",
    "PAZ Lush ice",
    "PAZ Lush ice +",
    "PAZ Mango",
    "PAZ Mango winter +",
    "PAZ Berry frost",
    "PAZ Berry frost +",
    "V&YOU Boost+ Cool Berry",
    "V&YOU Boost+ Intense Mint",
    "V&YOU Boost+ Fresh Citrus",
    "V&YOU Boost+ Blueberry Ice",
    "V&YOU Boost+ Spearmint",
    "V&YOU Boost+ Grape soda",
    "V&YOU Boost+ Berry Fizz",
    "V&YOU Boost+ Mint Freeze",
    "V&YOU Boost+ Berry Kiwi rush",
    "V&YOU Boost+ Tropical fusion",
    "V&YOU Boost Max Frostbite",
    "V&YOU Boost Max Frutti Blast",
    "V&YOU Boost Max Savage mango",
    "01 Riot Kit Cherry cola",
    "02 Riot Kit Pink lemonade",
    "03 Riot Kit Mango peach pineapple",
    "04 Riot Kit Grape ice",
    "05 Riot Kit Blueberry sour raspberry",
    "01 Riot Capsule Cherry cola",
    "02 Riot Capsule Pink lemonade",
    "03 Riot Capsule Mango peach pineapple",
    "04 Riot Capsule Grape ice",
    "05 Riot Capsule Blueberry sour raspberry",
    "06 Riot Capsule Strawberry blueberry ice",
    "07 Riot Capsule Blue cherry burst",
    "08 Riot Capsule Classic tobacoo",
    "09 Riot Capsule Banana Ice",
    "10 Riot Capsule Lime",
    "11 Riot Capsule Strawberry Kiwi Apple",
    "12 Riot Capsule Triple Mint",
]

BRAND_BOUNDARIES = {
    "PAZ Berry frost +": "PAZ",
    "V&YOU Boost Max Savage mango": "V&YOU",
    "05 Riot Kit Blueberry sour raspberry": "RIOT Kit",
    "12 Riot Capsule Triple Mint": "RIOT Capsule",
}

BRAND_PRODUCTS = {
    "PAZ": PRODUCT_ORDER[0:10],
    "V&YOU": PRODUCT_ORDER[10:23],
    "RIOT Kit": PRODUCT_ORDER[23:28],
    "RIOT Capsule": PRODUCT_ORDER[28:40],
}


class TextExtractor(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.chunks: List[str] = []

    def handle_data(self, data: str) -> None:
        data = data.strip()
        if data:
            self.chunks.append(data)


def html_to_chunks(html: str) -> List[str]:
    parser = TextExtractor()
    try:
        parser.feed(html)
        parser.close()
    except Exception:
        pass
    return parser.chunks


def extract_predal_from_chunks(chunks: List[str]) -> Optional[str]:
    upper_chunks = [c.upper() for c in chunks]
    for i, chunk_upper in enumerate(upper_chunks):
        if "ПРЕДАЛ" in chunk_upper:
            original = chunks[i]
            idx = original.upper().find("ПРЕДАЛ")
            after = original[idx + len("ПРЕДАЛ"):]
            after = after.lstrip(" :\u00a0")
            if after:
                return after
            if i + 1 < len(chunks):
                next_chunk = chunks[i + 1].strip()
                if next_chunk and not next_chunk.endswith(":"):
                    return next_chunk
    return None


def clean_predal(value: str) -> str:
    if not value:
        return ""
    text = value.strip()
    marker = "(име, фамилия, подпис):"
    lower = text.lower()
    idx = lower.find(marker)
    if idx != -1:
        text = (text[:idx] + text[idx + len(marker):]).strip()
    return text.strip(" :\u00a0")


def fetch_html(url: str, timeout: float) -> Tuple[str, str]:
    headers = {"User-Agent": UA}
    resp = requests.get(url, headers=headers, timeout=timeout)
    resp.raise_for_status()
    content_type = resp.headers.get("Content-Type", "")
    resp.encoding = resp.apparent_encoding or resp.encoding
    return resp.text, content_type


def build_header_map(ws) -> Dict[str, int]:
    header_map: Dict[str, int] = {}
    for cell in ws[1]:
        if cell.value is None:
            continue
        header_map[str(cell.value).strip()] = cell.column
    return header_map


def format_cell(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    return str(value).strip()


def load_rows(
    xlsx_path: str,
    client_prefix: str,
    number_header: str = "Номер",
    client_header: str = "Клиент",
    date_header: str = "Дата на документа",
    product_header: str = "Наименование на продукта",
    quantity_header: str = "Количество",
) -> List[Tuple[str, str, str, str, str, Optional[str]]]:
    wb = load_workbook(xlsx_path)
    ws = wb.active

    header_map = build_header_map(ws)

    def col(name: str, default: int) -> int:
        return header_map.get(name, default)

    number_col = col(number_header, 2)
    client_col = col(client_header, 4)
    date_col = col(date_header, 6)
    product_col = col(product_header, 8)
    quantity_col = col(quantity_header, 10)

    rows: List[Tuple[str, str, str, str, str, Optional[str]]] = []
    prefix = client_prefix.strip().lower()

    for r in range(2, ws.max_row + 1):
        client = ws.cell(row=r, column=client_col).value
        if client is None:
            continue
        client_str = str(client).strip()
        if not client_str.lower().startswith(prefix):
            continue

        number_cell = ws.cell(row=r, column=number_col)
        number = number_cell.value
        if number is None:
            continue
        number_str = format_cell(number)

        date_str = format_cell(ws.cell(row=r, column=date_col).value)
        product_str = format_cell(ws.cell(row=r, column=product_col).value)
        qty_str = format_cell(ws.cell(row=r, column=quantity_col).value)

        link = None
        if number_cell.hyperlink and number_cell.hyperlink.target:
            link = number_cell.hyperlink.target
        elif isinstance(number_cell.value, str) and number_cell.value.startswith("http"):
            link = number_cell.value

        rows.append((number_str, client_str, date_str, product_str, qty_str, link))

    return rows


def build_output(
    output_path: str, rows: List[Tuple[str, str, str, str, str, Optional[str]]]
) -> None:
    wb = Workbook()
    ws_default = wb.active
    brand_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")

    def safe_title(name: str, existing: set) -> str:
        invalid = set('[]:*?/\\')
        base = "".join("_" if c in invalid else c for c in name).strip()
        if not base:
            base = "Неизвестен"
        base = base[:31]
        title = base
        counter = 2
        while title in existing:
            suffix = f" {counter}"
            title = (base[: 31 - len(suffix)] + suffix)[:31]
            counter += 1
        existing.add(title)
        return title

    def autosize_worksheet(ws) -> None:
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value is None:
                    continue
                val = str(cell.value)
                if len(val) > max_len:
                    max_len = len(val)
            if max_len:
                ws.column_dimensions[col_letter].width = min(60, max_len + 2)

    def append_row_with_link(
        ws,
        row_data: Tuple[str, str, str, str, str, Optional[str]],
    ) -> None:
        number, predal, date_str, product, qty, link = row_data
        ws.append([number, predal, date_str, product, qty])
        if link:
            cell = ws.cell(row=ws.max_row, column=1)
            cell.hyperlink = link
            cell.style = "Hyperlink"

    # Preserve order of appearance for colleagues
    colleague_order: List[str] = []
    by_colleague: Dict[str, List[Tuple[str, str, str, str, str, Optional[str]]]] = {}
    for row in rows:
        name = row[1].strip() if row[1].strip() else "Неизвестен"
        if name not in by_colleague:
            by_colleague[name] = []
            colleague_order.append(name)
        by_colleague[name].append(row)

    existing_titles = set()
    for name in colleague_order:
        ws = wb.create_sheet(title=safe_title(name, existing_titles))
        ws.append(["Номер", "Предал", "Дата", "Продукт", "Количество"])
        for row in by_colleague[name]:
            append_row_with_link(ws, row)

        date_headers, summary_rows = build_summary_for_colleague(by_colleague[name])
        ws.append([])
        ws.append(["Обобщение по продукт"])
        ws.append(["Продукт", *date_headers])
        for product, totals in summary_rows:
            ws.append([product, *totals])
            if product == "":
                row_idx = ws.max_row
                for col_idx in range(1, len(totals) + 2):
                    ws.cell(row=row_idx, column=col_idx).fill = brand_fill
        autosize_worksheet(ws)

    # Remove default empty sheet
    wb.remove(ws_default)
    wb.save(output_path)


def parse_quantity(value: str) -> Optional[float]:
    text = value.strip()
    if not text:
        return None
    text = text.replace(" ", "")
    if "," in text and "." in text:
        text = text.replace(",", "")
    elif "," in text:
        text = text.replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return None


def normalize_product(value: str) -> str:
    text = value.strip().lower()
    while "  " in text:
        text = text.replace("  ", " ")
    return text


def build_summary_for_colleague(
    rows: List[Tuple[str, str, str, str, str, Optional[str]]]
) -> Tuple[List[str], List[Tuple[str, List[float]]]]:
    product_map = {normalize_product(p): p for p in PRODUCT_ORDER}
    totals: Dict[Tuple[str, str], float] = {}
    date_set: set = set()

    for _num, _predal, date_str, product, qty, _link in rows:
        if date_str:
            date_set.add(str(date_str))
        prod_key = normalize_product(product)
        if prod_key not in product_map:
            continue
        prod = product_map[prod_key]
        amount = parse_quantity(qty)
        if amount is None:
            continue
        key = (prod, str(date_str) if date_str else "")
        totals[key] = totals.get(key, 0.0) + amount

    date_headers = sorted(date_set)
    if not date_headers:
        date_headers = ["Дата"]

    def format_num(value: float):
        return int(value) if value.is_integer() else value

    def totals_for_products(products: List[str]) -> List[float]:
        sums: List[float] = []
        for date_val in date_headers:
            sums.append(sum(totals.get((p, date_val), 0.0) for p in products))
        return sums

    summary_rows: List[Tuple[str, List[float]]] = []
    for prod in PRODUCT_ORDER:
        row_totals: List[float] = []
        for date_val in date_headers:
            val = totals.get((prod, date_val), 0.0)
            row_totals.append(format_num(val))
        summary_rows.append((prod, row_totals))

        if prod in BRAND_BOUNDARIES:
            brand = BRAND_BOUNDARIES[prod]
            brand_totals = totals_for_products(BRAND_PRODUCTS[brand])
            summary_rows.append(("", [format_num(v) for v in brand_totals]))

    return date_headers, summary_rows


class Worker(QtCore.QObject):
    progress = QtCore.pyqtSignal(int, int)
    done = QtCore.pyqtSignal(str, int)
    error = QtCore.pyqtSignal(str)

    def __init__(
        self,
        input_path: str,
        output_path: str,
        client_prefix: str,
        timeout: float,
        delay: float,
    ) -> None:
        super().__init__()
        self.input_path = input_path
        self.output_path = output_path
        self.client_prefix = client_prefix
        self.timeout = timeout
        self.delay = delay

    @QtCore.pyqtSlot()
    def run(self) -> None:
        try:
            rows = load_rows(self.input_path, self.client_prefix)
            if not rows:
                self.error.emit("Няма редове за този клиент.")
                return

            url_cache: Dict[str, Optional[str]] = {}
            output_rows: List[Tuple[str, str, str, str, str, Optional[str]]] = []
            total = len(rows)

            for i, (number, _client, date_str, product_str, qty_str, link) in enumerate(
                rows, start=1
            ):
                predal = ""
                if link:
                    if link in url_cache:
                        predal = url_cache[link] or ""
                    else:
                        try:
                            html, content_type = fetch_html(link, self.timeout)
                            if "text/html" in content_type:
                                chunks = html_to_chunks(html)
                                predal = extract_predal_from_chunks(chunks) or ""
                            predal = clean_predal(predal)
                            url_cache[link] = predal
                        except Exception:
                            url_cache[link] = ""
                    if self.delay:
                        time.sleep(self.delay)

                output_rows.append((number, predal, date_str, product_str, qty_str, link))
                self.progress.emit(i, total)

            build_output(self.output_path, output_rows)
            missing = sum(1 for _n, p, _d, _pr, _q, _l in output_rows if not p)
            self.done.emit(self.output_path, missing)
        except Exception as exc:
            self.error.emit(str(exc))


class MainWindow(QtWidgets.QWidget):
    def __init__(self) -> None:
        super().__init__()
        self.setWindowTitle("Лагардер – Стокови разписки")
        self.setFixedSize(620, 300)

        self.input_path = ""
        self.thread: Optional[QtCore.QThread] = None
        self.worker: Optional[Worker] = None

        self._build_ui()

    def _build_ui(self) -> None:
        layout = QtWidgets.QVBoxLayout(self)
        layout.setContentsMargins(16, 16, 16, 16)
        layout.setSpacing(10)

        file_row = QtWidgets.QHBoxLayout()
        self.file_edit = QtWidgets.QLineEdit()
        self.file_edit.setReadOnly(True)
        self.file_btn = QtWidgets.QPushButton("Избери…")
        self.file_btn.clicked.connect(self.select_file)
        file_row.addWidget(self.file_edit)
        file_row.addWidget(self.file_btn)
        layout.addWidget(QtWidgets.QLabel("Файл (.xlsx)"))
        layout.addLayout(file_row)

        layout.addWidget(QtWidgets.QLabel("Фирма започва с"))
        self.prefix_edit = QtWidgets.QLineEdit("Лагардер")
        layout.addWidget(self.prefix_edit)

        # Hidden advanced settings (fixed defaults)
        self.timeout_value = 20.0
        self.delay_value = 0.0

        self.run_btn = QtWidgets.QPushButton("Обработи")
        self.run_btn.clicked.connect(self.start)
        layout.addWidget(self.run_btn)

        self.progress = QtWidgets.QProgressBar()
        self.progress.setRange(0, 100)
        layout.addWidget(self.progress)

        self.status = QtWidgets.QLabel("")
        layout.addWidget(self.status)

    def select_file(self) -> None:
        path, _ = QtWidgets.QFileDialog.getOpenFileName(
            self, "Избери Excel файл", "", "Excel Files (*.xlsx)"
        )
        if path:
            self.input_path = path
            self.file_edit.setText(path)

    def start(self) -> None:
        if not self.input_path:
            QtWidgets.QMessageBox.critical(self, "Грешка", "Избери Excel файл.")
            return

        default_name = "lagardere_output.xlsx"
        out_path, _ = QtWidgets.QFileDialog.getSaveFileName(
            self,
            "Запиши резултата",
            os.path.join(os.path.dirname(self.input_path), default_name),
            "Excel Files (*.xlsx)",
        )
        if not out_path:
            return

        self.run_btn.setEnabled(False)
        self.progress.setValue(0)
        self.status.setText("Започвам...")

        self.thread = QtCore.QThread()
        self.worker = Worker(
            self.input_path,
            out_path,
            self.prefix_edit.text(),
            float(self.timeout_value),
            float(self.delay_value),
        )
        self.worker.moveToThread(self.thread)

        self.thread.started.connect(self.worker.run)
        self.worker.progress.connect(self.on_progress)
        self.worker.done.connect(self.on_done)
        self.worker.error.connect(self.on_error)
        self.worker.done.connect(self.thread.quit)
        self.worker.error.connect(self.thread.quit)
        self.thread.finished.connect(self.thread.deleteLater)

        self.thread.start()

    @QtCore.pyqtSlot(int, int)
    def on_progress(self, current: int, total: int) -> None:
        if total:
            self.progress.setValue(int((current / total) * 100))
        self.status.setText(f"Обработени: {current}/{total}")

    @QtCore.pyqtSlot(str, int)
    def on_done(self, output_path: str, missing: int) -> None:
        self.run_btn.setEnabled(True)
        self.progress.setValue(100)
        self.status.setText(f"Готово. Липсва 'Предал' за {missing} реда.")
        QtWidgets.QMessageBox.information(self, "Готово", f"Файлът е записан: {output_path}")
        QtGui.QDesktopServices.openUrl(QtCore.QUrl.fromLocalFile(output_path))

    @QtCore.pyqtSlot(str)
    def on_error(self, message: str) -> None:
        self.run_btn.setEnabled(True)
        self.status.setText("Грешка")
        QtWidgets.QMessageBox.critical(self, "Грешка", message)



def main() -> None:
    app = QtWidgets.QApplication([])
    app.setApplicationName("Lagardere")
    window = MainWindow()
    window.show()
    app.exec()


if __name__ == "__main__":
    main()
