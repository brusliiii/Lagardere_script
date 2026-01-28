#!/usr/bin/env python3
"""Extract 'ПРЕДАЛ:' names from linked documents in an Excel file.

Reads hyperlinks from the 'Номер' column (column B by default), opens each
linked document, parses the name after 'ПРЕДАЛ:', and writes a two-column CSV.
"""

import argparse
import csv
import sys
import time
import urllib.request
from html.parser import HTMLParser
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook


UA = "Mozilla/5.0 (compatible; extract_predal/1.0)"


class TextExtractor(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.chunks: List[str] = []

    def handle_data(self, data: str) -> None:
        data = data.strip()
        if data:
            self.chunks.append(data)


def extract_predal_from_chunks(chunks: List[str]) -> Optional[str]:
    upper_chunks = [c.upper() for c in chunks]
    for i, chunk_upper in enumerate(upper_chunks):
        if "ПРЕДАЛ" in chunk_upper:
            original = chunks[i]
            # Try to extract from same chunk (e.g. "ПРЕДАЛ: Иван Иванов")
            idx = original.upper().find("ПРЕДАЛ")
            after = original[idx + len("ПРЕДАЛ"):]
            after = after.lstrip(" :\u00a0")
            if after:
                return after
            # Otherwise, look at the next chunk
            if i + 1 < len(chunks):
                next_chunk = chunks[i + 1].strip()
                if next_chunk and not next_chunk.endswith(":"):
                    return next_chunk
    return None


def html_to_chunks(html: str) -> List[str]:
    parser = TextExtractor()
    try:
        parser.feed(html)
        parser.close()
    except Exception:
        # Best-effort parsing
        pass
    return parser.chunks


def fetch_html(url: str, cookie: Optional[str], timeout: float) -> Tuple[str, str]:
    headers = {"User-Agent": UA}
    if cookie:
        headers["Cookie"] = cookie
    req = urllib.request.Request(url, headers=headers)
    with urllib.request.urlopen(req, timeout=timeout) as resp:
        content_type = resp.headers.get("Content-Type", "")
        charset = resp.headers.get_content_charset() or "utf-8"
        raw = resp.read()
        try:
            text = raw.decode(charset, errors="replace")
        except LookupError:
            text = raw.decode("utf-8", errors="replace")
    return text, content_type


def load_links(
    xlsx_path: str,
    sheet: Optional[str],
    number_header: str,
) -> List[Tuple[str, str]]:
    wb = load_workbook(xlsx_path)
    ws = wb[sheet] if sheet else wb.active

    header_row = 1
    header_map = {}
    for cell in ws[header_row]:
        if cell.value:
            header_map[str(cell.value).strip()] = cell.column

    number_col = header_map.get(number_header, 2)  # default column B

    seen = set()
    results: List[Tuple[str, str]] = []

    for row in range(header_row + 1, ws.max_row + 1):
        cell = ws.cell(row=row, column=number_col)
        doc_number = cell.value
        if doc_number is None:
            continue
        doc_number = str(doc_number).strip()
        if not doc_number or doc_number in seen:
            continue

        link = None
        if cell.hyperlink and cell.hyperlink.target:
            link = cell.hyperlink.target
        elif isinstance(cell.value, str) and cell.value.startswith("http"):
            link = cell.value

        if link:
            results.append((doc_number, link))
            seen.add(doc_number)

    return results


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Extract 'ПРЕДАЛ:' from hyperlinks in an Excel file."
    )
    parser.add_argument("input", help="Input .xlsx file")
    parser.add_argument("output", help="Output CSV file")
    parser.add_argument(
        "--sheet",
        default=None,
        help="Sheet name (default: active sheet)",
    )
    parser.add_argument(
        "--number-header",
        default="Номер",
        help="Header name for the number column (default: 'Номер')",
    )
    parser.add_argument(
        "--cookie",
        default=None,
        help="Cookie header string if login is required",
    )
    parser.add_argument(
        "--timeout",
        type=float,
        default=20.0,
        help="Request timeout in seconds",
    )
    parser.add_argument(
        "--delay",
        type=float,
        default=0.0,
        help="Delay between requests in seconds",
    )
    parser.add_argument(
        "--include-link",
        action="store_true",
        help="Include the document link in output",
    )
    args = parser.parse_args()

    try:
        links = load_links(args.input, args.sheet, args.number_header)
    except Exception as exc:
        print(f"Failed to read Excel file: {exc}", file=sys.stderr)
        return 1

    if not links:
        print("No hyperlinks found in the number column.", file=sys.stderr)
        return 1

    url_cache: Dict[str, Optional[str]] = {}
    results: List[Tuple[str, Optional[str], str]] = []

    for doc_number, url in links:
        if url in url_cache:
            predal = url_cache[url]
        else:
            try:
                html, content_type = fetch_html(url, args.cookie, args.timeout)
                if "text/html" not in content_type:
                    predal = None
                else:
                    chunks = html_to_chunks(html)
                    predal = extract_predal_from_chunks(chunks)
                url_cache[url] = predal
            except Exception as exc:
                print(f"Failed to fetch {url}: {exc}", file=sys.stderr)
                predal = None
                url_cache[url] = predal

        results.append((doc_number, predal, url))
        if args.delay:
            time.sleep(args.delay)

    with open(args.output, "w", newline="", encoding="utf-8") as out_f:
        writer = csv.writer(out_f)
        header = ["Номер", "Предал"]
        if args.include_link:
            header.append("Линк")
        writer.writerow(header)
        for doc_number, predal, url in results:
            row = [doc_number, predal or ""]
            if args.include_link:
                row.append(url)
            writer.writerow(row)

    missing = sum(1 for _, predal, _ in results if not predal)
    print(
        f"Processed {len(results)} documents. Missing 'Предал' for {missing} docs."
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
