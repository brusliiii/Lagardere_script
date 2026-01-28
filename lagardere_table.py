#!/usr/bin/env python3
"""Build a table for stock receipts filtered by client prefix.

Output columns:
A: Номер
B: Предал (from document hyperlink)
C: Дата
D: Продукт
E: Количество
"""

import argparse
import os
import sys
import time
import urllib.request
from html.parser import HTMLParser
from typing import Dict, List, Optional, Tuple

from openpyxl import Workbook, load_workbook

UA = "Mozilla/5.0 (compatible; lagardere_table/1.0)"


class TextExtractor(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.chunks: List[str] = []

    def handle_data(self, data: str) -> None:
        data = data.strip()
        if data:
            self.chunks.append(data)


def extract_predal_from_chunks(chunks: List[str]) -> Optional[str]:
    upper_chunks = [c.upper() for c in chunks]
    for i, chunk_upper in enumerate(upper_chunks):
        if "ПРЕДАЛ" in chunk_upper:
            original = chunks[i]
            idx = original.upper().find("ПРЕДАЛ")
            after = original[idx + len("ПРЕДАЛ"):]
            after = after.lstrip(" :\u00a0")
            if after:
                return after
            if i + 1 < len(chunks):
                next_chunk = chunks[i + 1].strip()
                if next_chunk and not next_chunk.endswith(":"):
                    return next_chunk
    return None


def html_to_chunks(html: str) -> List[str]:
    parser = TextExtractor()
    try:
        parser.feed(html)
        parser.close()
    except Exception:
        pass
    return parser.chunks


def fetch_html(url: str, cookie: Optional[str], timeout: float) -> Tuple[str, str]:
    headers = {"User-Agent": UA}
    if cookie:
        headers["Cookie"] = cookie
    req = urllib.request.Request(url, headers=headers)
    with urllib.request.urlopen(req, timeout=timeout) as resp:
        content_type = resp.headers.get("Content-Type", "")
        charset = resp.headers.get_content_charset() or "utf-8"
        raw = resp.read()
        try:
            text = raw.decode(charset, errors="replace")
        except LookupError:
            text = raw.decode("utf-8", errors="replace")
    return text, content_type


def build_header_map(ws) -> Dict[str, int]:
    header_map: Dict[str, int] = {}
    for cell in ws[1]:
        if cell.value is None:
            continue
        header_map[str(cell.value).strip()] = cell.column
    return header_map


def load_rows(
    xlsx_path: str,
    sheet: Optional[str],
    number_header: str,
    client_header: str,
    date_header: str,
    product_header: str,
    quantity_header: str,
    client_prefix: str,
) -> List[Tuple[str, str, str, str, str, Optional[str]]]:
    wb = load_workbook(xlsx_path)
    ws = wb[sheet] if sheet else wb.active

    header_map = build_header_map(ws)

    def col(name: str, default: int) -> int:
        return header_map.get(name, default)

    number_col = col(number_header, 2)
    client_col = col(client_header, 4)
    date_col = col(date_header, 6)
    product_col = col(product_header, 8)
    quantity_col = col(quantity_header, 10)

    rows: List[Tuple[str, str, str, str, str, Optional[str]]] = []
    prefix = client_prefix.strip().lower()

    for r in range(2, ws.max_row + 1):
        client = ws.cell(row=r, column=client_col).value
        if client is None:
            continue
        client_str = str(client).strip()
        if not client_str.lower().startswith(prefix):
            continue

        number_cell = ws.cell(row=r, column=number_col)
        number = number_cell.value
        if number is None:
            continue
        number_str = str(number).strip()

        date_val = ws.cell(row=r, column=date_col).value
        date_str = "" if date_val is None else str(date_val).strip()

        product_val = ws.cell(row=r, column=product_col).value
        product_str = "" if product_val is None else str(product_val).strip()

        qty_val = ws.cell(row=r, column=quantity_col).value
        qty_str = "" if qty_val is None else str(qty_val).strip()

        link = None
        if number_cell.hyperlink and number_cell.hyperlink.target:
            link = number_cell.hyperlink.target
        elif isinstance(number_cell.value, str) and number_cell.value.startswith("http"):
            link = number_cell.value

        rows.append((number_str, client_str, date_str, product_str, qty_str, link))

    return rows


def write_output(output_path: str, rows: List[Tuple[str, str, str, str, str]]) -> None:
    if output_path.lower().endswith(".xlsx"):
        wb = Workbook()
        ws = wb.active
        ws.title = "Lagardere"
        ws.append(["Номер", "Предал", "Дата", "Продукт", "Количество"])
        for row in rows:
            ws.append(list(row))
        wb.save(output_path)
    else:
        import csv

        with open(output_path, "w", newline="", encoding="utf-8") as out_f:
            writer = csv.writer(out_f)
            writer.writerow(["Номер", "Предал", "Дата", "Продукт", "Количество"])
            writer.writerows(rows)


def main() -> int:
    parser = argparse.ArgumentParser(
        description=(
            "Build a table for clients starting with a prefix, using hyperlinks "
            "to fetch the 'ПРЕДАЛ' name."
        )
    )
    parser.add_argument("input", help="Input .xlsx file")
    parser.add_argument("output", help="Output CSV file")
    parser.add_argument("--sheet", default=None, help="Sheet name (default: active)")
    parser.add_argument("--client-prefix", default="Лагардер", help="Client prefix")
    parser.add_argument("--number-header", default="Номер", help="Number column header")
    parser.add_argument("--client-header", default="Клиент", help="Client column header")
    parser.add_argument("--date-header", default="Дата на документа", help="Date column header")
    parser.add_argument(
        "--product-header",
        default="Наименование на продукта",
        help="Product column header",
    )
    parser.add_argument("--quantity-header", default="Количество", help="Quantity header")
    parser.add_argument("--cookie", default=None, help="Cookie header string")
    parser.add_argument(
        "--cookie-file",
        default=None,
        help="Path to a text file containing the Cookie header value",
    )
    parser.add_argument(
        "--cookie-env",
        default=None,
        help="Environment variable name holding the Cookie header value",
    )
    parser.add_argument("--timeout", type=float, default=20.0, help="Request timeout")
    parser.add_argument("--delay", type=float, default=0.0, help="Delay between requests")

    args = parser.parse_args()

    cookie = args.cookie
    if not cookie and args.cookie_file:
        try:
            with open(args.cookie_file, "r", encoding="utf-8") as f:
                cookie = f.read().strip()
        except OSError as exc:
            print(f"Failed to read cookie file: {exc}", file=sys.stderr)
            return 1
    if not cookie and args.cookie_env:
        cookie = os.environ.get(args.cookie_env)

    try:
        rows = load_rows(
            args.input,
            args.sheet,
            args.number_header,
            args.client_header,
            args.date_header,
            args.product_header,
            args.quantity_header,
            args.client_prefix,
        )
    except Exception as exc:
        print(f"Failed to read Excel file: {exc}", file=sys.stderr)
        return 1

    if not rows:
        print("No matching rows for the client prefix.", file=sys.stderr)
        return 1

    url_cache: Dict[str, Optional[str]] = {}
    output_rows = []

    for number, _client, date_str, product_str, qty_str, link in rows:
        predal = None
        if link:
            if link in url_cache:
                predal = url_cache[link]
            else:
                try:
                    html, content_type = fetch_html(link, cookie, args.timeout)
                    if "text/html" in content_type:
                        chunks = html_to_chunks(html)
                        predal = extract_predal_from_chunks(chunks)
                    url_cache[link] = predal
                except Exception as exc:
                    print(f"Failed to fetch {link}: {exc}", file=sys.stderr)
                    url_cache[link] = None
            if args.delay:
                time.sleep(args.delay)

        output_rows.append((number, predal or "", date_str, product_str, qty_str))

    write_output(args.output, output_rows)

    missing = sum(1 for _n, p, _d, _pr, _q in output_rows if not p)
    print(f"Wrote {len(output_rows)} rows. Missing 'Предал' for {missing} rows.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
